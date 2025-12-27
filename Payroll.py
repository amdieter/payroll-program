import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import os
import pandas as pd
import xlwings
import tkinter as tk
from tkinter.ttk import Progressbar

# Program

def payroll():
    run_button.pack_forget()

    import time 
    progress.pack()
    progress['value'] = 20
    master.update_idletasks()

    shifts_wb = openpyxl.load_workbook('shifts data.xlsx')
    shifts = shifts_wb['shifts data']
    f_tips = shifts_wb['flight tips']
    twr_tips = shifts_wb['twr tips']

    # Sort shifts

    # Creates 'Name' column
    shifts.insert_cols(4)
    shifts['D1'] = "Name"

    # shifts to DataFrame
    shifts_df = pd.DataFrame(shifts.values)

    # Concat first and last names into Name
    for i in range(1, len(shifts_df)-1):
        shifts_df.loc[i,3] = shifts_df.loc[i,1] + shifts_df.loc[i,2]

    # Shift column names, sort
    shifts_df.columns = shifts_df.iloc[0]
    new_shifts_df = shifts_df[1:]
    new_shifts_df.index = new_shifts_df.index - 1
    sorted_shifts_df = new_shifts_df.sort_values(by=["Clockin date","Location","Job title","Name"], ignore_index=True)
    sorted_shifts_df

    # Replace the data in the sheet (first by clearing sheet)
    shifts.delete_rows(1,shifts.max_row)
    for i in dataframe_to_rows(sorted_shifts_df, index=False, header=True):
        shifts.append(i)

    # Reformating columns
    columns_to_shrink = ['A','B','C','I','J','K','L','M','N','O','P','Q','R','S','U','V','W','X','Y','Z','AA']
    for col in columns_to_shrink:
        shifts.column_dimensions[col].width = 1
    shifts.column_dimensions['D'].width = 13.33
    shifts.column_dimensions['G'].width = 16.73
    shifts.column_dimensions['T'].width = 12.67
    shifts.column_dimensions['AC'].width = 12.47
    shifts.column_dimensions['AD'].width = 10.73
    shifts.column_dimensions['AH'].width = 17

    # get current table row max for later
    shifts_row_max = shifts.max_row


    progress['value'] = 40
    master.update_idletasks()
    # Transpose tips for F and TWR and put in shifts sheet

    # flight tips
    # slice, create list, convert to df
    f_range = f_tips['A11':str(get_column_letter(f_tips.max_column))+'20']
    f_list = []
    for row in f_range:
        list = []
        for col in row:
            list.append(col.value)
        f_list.append(list)
    f_df = pd.DataFrame(f_list)

    # transpose
    transposed_f_df = f_df.T

    # change column names
    transposed_f_df.columns = transposed_f_df.iloc[0]
    new_transposed_f_df = transposed_f_df[1:]
    new_transposed_f_df.index = new_transposed_f_df.index - 1

    # slice to Sales and Tip
    sales_f_tip = new_transposed_f_df[["Sales","Tip"]]
    sales_f_tip.columns = ["Sales", "F Tips"]


    # TWR tips
    twr_range = twr_tips['A11':str(get_column_letter(twr_tips.max_column))+'20']
    twr_list = []
    for row in twr_range:
        list = []
        for col in row:
            list.append(col.value)
        twr_list.append(list)
    twr_df = pd.DataFrame(twr_list)
    transposed_twr_df = twr_df.T
    transposed_twr_df.columns = transposed_twr_df.iloc[0]
    new_transposed_twr_df = transposed_twr_df[1:]
    new_transposed_twr_df.index = new_transposed_twr_df.index - 1
    sales_twr_tip = new_transposed_twr_df[["Sales","Tip"]]
    sales_twr_tip.columns = ["Sales","TWR Tips"]

    # Combine data
    full_tip_df = pd.merge(sales_f_tip, sales_twr_tip, on='Sales',how='outer')

    pd.set_option('future.no_silent_downcasting', True)
    full_tip_df.fillna(0,inplace=True)

    # Paste into shifts sheet
    for r_idx, row in enumerate(dataframe_to_rows(full_tip_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=34):
            shifts.cell(row=r_idx, column=c_idx, value=value)


    progress['value'] = 60
    master.update_idletasks()
    
    #  add these  Daily Tips	Distributed Tips	Adjusted Tips	Tips/hr

    shifts['AB1'] = 'Daily Tips'
    shifts['AC1'] = 'Distributed Tips'
    shifts['AD1'] = 'Adjusted Tips'
    shifts['AE1'] = 'Tips/hr'


    # filling Daily Tips for Flight and TWR

    count = 0
    tip_indexes = []     # use this for later loopage
    for date in full_tip_df["Sales"]: # loop through each date
        for i in range(2, shifts_row_max):
            if shifts.cell(i,6).value == 'Flight' and date == shifts.cell(i,7).value:    # check that it is twr and if the date is the same
                shifts['AB'+str(i)] = full_tip_df["F Tips"].iloc[count]   # adds respective tip value onto sheet
                tip_indexes.append(i)
                break
        for i in range(2, shifts_row_max):
            if shifts.cell(i,6).value == 'The Wine Reserve' and date == shifts.cell(i,7).value:
                shifts['AB'+str(i)] = full_tip_df["TWR Tips"].iloc[count]
                tip_indexes.append(i)
                break
        count += 1


    # filling Distributed Tips

    for idx in range(len(tip_indexes)):
        list = []
        if (idx+1) > (len(tip_indexes)-1):   # case for last index
            for i in range(tip_indexes[idx], shifts_row_max):
                list.append(i)
            for item in list:
                shifts['AC'+str(item)] = '=AB${}*T{}/SUM(T${}:T${})'.format(list[0],item,list[0],list[-1])    
            break
        for i in range(tip_indexes[idx], tip_indexes[idx+1]):
            list.append(i)
        for item in list:
            shifts['AC'+str(item)] = '=AB${}*T{}/SUM(T${}:T${})'.format(list[0],item,list[0],list[-1])


    # filling Adjusted Tips

    for idx in range(len(tip_indexes)):
        kitchen_list = []
        bartender_list = []
        full_list = []
        if (idx+1) > (len(tip_indexes)-1):   # case for last index
            for i in range(tip_indexes[idx], shifts_row_max):
                if shifts['E'+str(i)].value == 'Kitchen':
                    kitchen_list.append(i)
                else:
                    bartender_list.append(i)
                full_list.append(i)
            for i in kitchen_list:
                shifts['AD'+str(i)] = '=AC{}/4'.format(str(i))
            for i in bartender_list:
                if len(kitchen_list) == 0:
                    shifts['AD'+str(i)] = '=AC{}'.format(i)
                else:
                    shifts['AD'+str(i)] = '=(AB${}-SUM(AD${}:AD${}))*T{}/(SUM(T${}:T${}))'.format(full_list[0],kitchen_list[0],kitchen_list[-1],i,bartender_list[0],bartender_list[-1])
            break
            
        for i in range(tip_indexes[idx], tip_indexes[idx+1]):
            if shifts['E'+str(i)].value == 'Kitchen':
                kitchen_list.append(i)
            else:
                bartender_list.append(i)
            full_list.append(i)
        for i in kitchen_list:
            shifts['AD'+str(i)] = '=AC{}/4'.format(str(i))
        for i in bartender_list:
            if len(kitchen_list) == 0:
                shifts['AD'+str(i)] = '=AC{}'.format(i)
            else:
                shifts['AD'+str(i)] = '=(AB${}-SUM(AD${}:AD${}))*T{}/(SUM(T${}:T${}))'.format(full_list[0],kitchen_list[0],kitchen_list[-1],i,bartender_list[0],bartender_list[-1])


    # filling Tips/hr

    for idx in range(len(tip_indexes)):
        list = []
        if (idx+1) > (len(tip_indexes)-1):   # case for last index
            for i in range(tip_indexes[idx], shifts_row_max):
                list.append(i)
            for i in list:
                shifts['AE'+str(i)] = '=AD{}/T{}'.format(i,i)  
            break
        for i in range(tip_indexes[idx], tip_indexes[idx+1]):
            list.append(i)
        for i in list:
            shifts['AE'+str(i)] = '=AD{}/T{}'.format(i,i)


    # total Daily Tips	Distributed Tips	Adjusted Tips

    shifts['AB'+str(shifts_row_max)] = '=SUM(AB2:AB{})'.format(shifts_row_max-1)
    shifts['AC'+str(shifts_row_max)] = '=SUM(AC2:AC{})'.format(shifts_row_max-1)
    shifts['AD'+str(shifts_row_max)] = '=SUM(AD2:AD{})'.format(shifts_row_max-1)



    # Change columns in 1data to currency format

    for col in ['AB','AC','AD','AI','AJ']:
        for i in range(2,shifts_row_max+1):
            shifts[col+str(i)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


    # Save to this file
    shifts_wb.save('1data.xlsx')

    progress['value'] = 80
    master.update_idletasks()
    # create dataframe with Name, Hours, Adj. Tips, and Tips/hr. to get unique and sumif

    # dataframe with values instead of formulas
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open('1data.xlsx')
    excel_book.save()
    excel_book.close()
    excel_app.quit()

    # load shifts data that we will save into new file to get real values instead of formulas
    shifts_wb = openpyxl.load_workbook('1data.xlsx', data_only = True)
    shifts = shifts_wb['shifts data']

    # Make summary sheet
    summary = shifts_wb.create_sheet("Summary")

    # Reformat dataframe
    fin_shifts_df = pd.DataFrame(shifts.values)
    fin_shifts_df.columns = fin_shifts_df.iloc[0]
    fin_shifts_df = fin_shifts_df[1:]
    fin_shifts_df.index = fin_shifts_df.index - 1
    nhat_shifts_df = fin_shifts_df[["Name","Job title","Total paid hours","Adjusted Tips"]]
    sliced_shifts_df = nhat_shifts_df.head(shifts_row_max-2)

    # Groupby and aggregate
    unique_shifts_df = sliced_shifts_df.groupby(sliced_shifts_df["Name"]).aggregate({'Job title': 'first','Total paid hours': 'sum', 'Adjusted Tips': 'sum'})
    unique_shifts_df = unique_shifts_df.reset_index()

    # paste dataframe into Summary
    for i in dataframe_to_rows(unique_shifts_df, index=False, header=True):
        summary.append(i)


    # Add additional columns
    summary['E1'] = 'Tips/hr'
    summary['F1'] = 'Additional'
    summary['G1'] = 'Adj. Tips/hr'

    # Filling Tips/hr
    for i in range(2, (summary.max_row + 1)):
        summary['E'+str(i)] = summary['D'+str(i)].value/summary['C'+str(i)].value

    # Filling Additional
    for i in range(2, (summary.max_row + 1)):
        if summary['B'+str(i)].value == 'Kitchen':
            if summary['E'+str(i)].value > 3:    # base pay for kitchen
                summary['F'+str(i)] = 0
            else:
                summary['F'+str(i)] = '=(3-E{})*C{}'.format(i,i)
        else:
            if summary['E'+str(i)].value > 10:     # base pay for bartenders
                summary['F'+str(i)] = 0
            else:
                summary['F'+str(i)] = '=(10-E{})*C{}'.format(i,i)

    # Filling Adj. Tips/hr
    for i in range(2, (summary.max_row + 1)):
        summary['G'+str(i)] = '=(F{}+D{})/C{}'.format(i,i,i)

    progress['value'] = 100
    master.update_idletasks()

    # total Hours, Adj. tips, Additional
    summary_rows = (summary.max_row)
    summary['C'+str(summary_rows+1)] = '=SUM(C2:C{})'.format(summary_rows)
    summary['D'+str(summary_rows+1)] = '=SUM(D2:D{})'.format(summary_rows)
    summary['F'+str(summary_rows+1)] = '=SUM(F2:F{})'.format(summary_rows)


    # Copy Adj Tips and Additional and swap them

    # Additional
    i = 1
    for cell in summary['F']:
        summary['J'+str(i)] = cell.value
        i += 1

    # Adj. Tips
    j = 1
    for cell in summary['D']:
        summary['K'+str(j)] = cell.value
        j += 1


    # Change columns in 1Summary to currency format

    for col in ['D','F','J','K']:
        for i in range(2,summary_rows+2):
            summary[col+str(i)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Change dimensions to fit total
    summary.column_dimensions['D'].width = 10
    summary.column_dimensions['K'].width = 10

    # Delete sheets not needed in summary
    try:
        del shifts_wb['Sheet1']
    except KeyError:
        pass
    del shifts_wb['shifts data']
    del shifts_wb['flight tips']
    del shifts_wb['twr tips']

    # Save/Create new file Summary
    shifts_wb.save('1Summary.xlsx')

    progress.pack_forget()
    error_label.config(text="Payroll complete. Please exit.")





def submit():
    error_label.config(text="")
    folder_name = folder_entry_var.get()

    with open("saved_directory.txt", "r") as file:
        curr_dir = file.read().strip()

    path = curr_dir + ('/') + folder_name
    
    # test if the directory is valid
    try: 
        os.chdir(path)
        
        # Ensure it contains shifts data workbook
        try: 
            shifts_wb = openpyxl.load_workbook('shifts data.xlsx')

            # Check shifts data exists
            try:
                shifts = shifts_wb['shifts data']

                # Check that flight tips exists
                try:
                    f_tips = shifts_wb['flight tips']

                    # Check that twr tips exists
                    try:
                        twr_tips = shifts_wb['twr tips']
                        error_label.config(text="Success!", font=('Times New Roman', 18)) # everything is present
                        submit_button.pack_forget()
                        run()
                        
                    except KeyError:
                        error_label.config(text="twr tips data isn't in the workbook. Ensure twr tips is present.", font=('Times New Roman', 12))
                        
                except KeyError:
                    error_label.config(text="flight tips isn't in the workbook. Ensure flight tips is present.", font=('Times New Roman', 12))
                    
            except KeyError:
                error_label.config(text="shifts data isn't in the workbook. Ensure shifts data is present.", font=('Times New Roman', 12))
                
        except FileNotFoundError:
            error_label.config(text="shifts data.xlsx does not exist in this directory. Ensure shifts data.xlsx is in the folder.", font=('Times New Roman', 12))
            
    except FileNotFoundError:
        error_label.config(text="Directory not found, please try again.", font=('Times New Roman', 14))
    
    # reset the entry textbox
    folder_entry_var.set("")

def run():
    run_button.pack(side='left')

def modify_directory():
    mod_dir_error_label.config(text="")
    path = mod_dir_entry_var.get()

    if os.path.exists(path):
        with open("saved_directory.txt", "w") as file:
            file.write(path)
        mod_dir_error_label.config(text="Default Directory Successfully Modified.", font=('Times New Roman', 8))
        def_dir.config(text=path, font=('Times New Roman', 8))
        
    else:
        mod_dir_error_label.config(text="Directory not found, please try again.", font=('Times New Roman', 8))
    mod_dir_entry_var.set("")
    


# Create GUI

master = tk.Tk()

master.geometry("500x350")
master.title("Payroll")

curr_dir = None
if os.path.exists("saved_directory.txt"):
    with open("saved_directory.txt", "r") as file:
        curr_dir = file.read().strip()
        if curr_dir == '':
            curr_dir = 'No directory saved yet.'
else:
    curr_dir = 'No directory saved yet.'

def_directory_frame = tk.Frame(master)
def_directory_frame.pack()

# current default directory label
def_dir_label = tk.Label(def_directory_frame, text="Current Default Directory: ", font=('Times New Roman', 8))
def_dir_label.grid(row=0, column=0, padx=3, pady=1)

def_dir = tk.Label(def_directory_frame, text=curr_dir, font=('Times New Roman', 8))
def_dir.grid(row=0, column=1, padx=0, pady=1)

modify_frame = tk.Frame(master)
modify_frame.pack()

# modify default directory
mod_dir_label = tk.Label(modify_frame, text="Modify Default Directory:", font=('Times New Roman', 8))
mod_dir_label.grid(row=0, column=0, padx=5, pady=3)

mod_dir_entry_var = tk.StringVar()
mod_dir_entry = tk.Entry(modify_frame, textvariable = mod_dir_entry_var, font=('Times New Roman', 8))
mod_dir_entry.grid(row=0, column=1, padx=5, pady=3)

mod_dir_button = tk.Button(modify_frame, text = 'Enter', command = modify_directory, font=('Times New Roman',8))
mod_dir_button.grid(row=0, column=2, padx=5, pady=3)

# initialize modify directory label
mod_dir_error_label = tk.Label(master, text="")
mod_dir_error_label.pack()

# Enter folder name: label
folder_label = tk.Label(master, text="Enter folder name:", font=('Times New Roman', 18))
folder_label.pack()

# initialize folder_name
folder_name = None

# entry textbox
folder_entry_var = tk.StringVar()
folder_entry = tk.Entry(master, textvariable = folder_entry_var, font=('Times New Roman', 18))
folder_entry.pack()

# initialize error label
error_label = tk.Label(master, text="")
error_label.pack()

progress = Progressbar(master, orient='horizontal', length=100,mode='determinate')

# buttons frame
button_frame = tk.Frame(master)
button_frame.pack()

# Submit button
submit_button = tk.Button(button_frame, text = 'Submit', command = submit, font=('Times New Roman',18))
submit_button.pack(side='left',padx=10)

run_button = tk.Button(button_frame, text = 'Run', command = payroll, font=('Times New Roman', 18))

master.mainloop()

# pyinstaller --onefile --noconsole Payroll.py